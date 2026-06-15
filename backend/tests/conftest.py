"""
Shared test fixtures for VAT Analytics tests.
"""

import os
import sys
import tempfile

import pytest
import openpyxl

TEST_DIR = os.path.dirname(__file__)
BACKEND_ROOT = os.path.dirname(TEST_DIR)
TEST_DATA_PATH = os.path.join(TEST_DIR, "test_data.xlsx")

# Gør backend-roden importérbar (main, auth, audit_log) når pytest kører fra repo-roden.
sys.path.insert(0, BACKEND_ROOT)

# Peg auth-/audit-databaserne mod en midlertidig mappe ALLEREDE ved import, så
# import-tids-initialisering af main.py aldrig skriver i repoet. Sæt også en fast
# SECRET_KEY + tillad cookies over http (TestClient), før main importeres.
_TMP = tempfile.mkdtemp(prefix="vatanalytics-test-")
os.environ.setdefault("AUTH_DB_PATH", os.path.join(_TMP, "auth.db"))
os.environ.setdefault("AUDIT_DB_PATH", os.path.join(_TMP, "audit.db"))
os.environ.setdefault("SECRET_KEY", "test-secret-key-fixed")
os.environ.setdefault("SESSION_COOKIE_SECURE", "0")


def _create_test_excel():
    """Create a small test Excel file with sample transactions."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Bilagsnr", "Dato", "Konto", "Kontobeskrivelse", "Beskrivelse",
        "Debet", "Kredit", "Moms", "Momskode", "Momssats",
        "Leverandørnr", "Leverandør", "Fakturanr", "Valuta",
    ]
    ws.append(headers)

    rows = [
        ["B001", "2024-01-15", "4000", "Varekøb", "Køb af varer fra ABC",
         10000.00, 0, 2500.00, "I25", 25.0, "L001", "ABC Supplies", "F-2024-001", "DKK"],
        ["B002", "2024-01-20", "4000", "Varekøb", "Køb af materialer",
         5000.00, 0, 1250.00, "I25", 25.0, "L002", "XYZ Materials", "F-2024-002", "DKK"],
        ["B003", "2024-02-01", "1000", "Salg", "Salg til kunde K001",
         0, 20000.00, 5000.00, "U25", 25.0, "", "", "S-2024-001", "DKK"],
        ["B004", "2024-02-10", "4000", "Varekøb", "Kontorartikler",
         800.00, 0, 200.00, "I25", 25.0, "L003", "Office Pro", "F-2024-003", "DKK"],
        ["B005", "2024-02-15", "4000", "Varekøb", "IT udstyr",
         15000.00, 0, 3750.00, "I25", 25.0, "L001", "ABC Supplies", "F-2024-004", "DKK"],
        ["B006", "2024-03-01", "1000", "Salg", "Konsulentydelse",
         0, 50000.00, 12500.00, "U25", 25.0, "", "", "S-2024-002", "DKK"],
        ["B007", "2024-03-05", "4000", "Varekøb", "Reservedele",
         3200.00, 0, 800.00, "I25", 25.0, "L004", "Parts Inc", "F-2024-005", "DKK"],
        ["B008", "2024-03-10", "4100", "Fragt", "Fragt og forsendelse",
         1500.00, 0, 375.00, "I25", 25.0, "L005", "Fast Freight", "F-2024-006", "DKK"],
    ]
    for row in rows:
        ws.append(row)

    wb.save(TEST_DATA_PATH)


# Ensure the test Excel file exists before tests run
if not os.path.exists(TEST_DATA_PATH):
    _create_test_excel()


@pytest.fixture
def test_excel_path():
    """Path to the test Excel file with sample transactions."""
    if not os.path.exists(TEST_DATA_PATH):
        _create_test_excel()
    return TEST_DATA_PATH


@pytest.fixture
def empty_excel_path(tmp_path):
    """Path to an empty Excel file (header only, no data rows)."""
    path = str(tmp_path / "empty.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Bilagsnr", "Dato", "Konto", "Beskrivelse", "Debet", "Kredit"])
    wb.save(path)
    return path


@pytest.fixture
def english_columns_excel(tmp_path):
    """Excel file with English column names."""
    path = str(tmp_path / "english.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "transaction_id", "date", "account_id", "description",
        "debit_amount", "credit_amount", "vat_amount", "vat_code", "vat_rate",
        "supplier_id", "supplier_name", "invoice_number", "currency",
    ])
    ws.append([
        "T001", "2024-01-15", "4000", "Purchase supplies",
        10000.00, 0, 2500.00, "I25", 25.0,
        "S001", "ACME Corp", "INV-001", "DKK",
    ])
    wb.save(path)
    return path


@pytest.fixture
def client(tmp_path, monkeypatch):
    """FastAPI TestClient med en frisk auth-/audit-db pr. test (ingen følgning af
    redirects, så vi kan assert'e status-koderne)."""
    monkeypatch.setenv("AUTH_DB_PATH", str(tmp_path / "auth.db"))
    monkeypatch.setenv("AUDIT_DB_PATH", str(tmp_path / "audit.db"))
    monkeypatch.setenv("SESSION_COOKIE_SECURE", "0")

    import auth
    import audit_log
    auth.init_db()
    audit_log.init_audit_db()

    from fastapi.testclient import TestClient
    from main import app
    return TestClient(app, follow_redirects=False)
