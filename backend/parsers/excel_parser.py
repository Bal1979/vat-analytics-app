"""
Excel/CSV Parser for VAT Analytics.
Konverterer Excel/CSV data til det samme standardformat som SAF-T parseren,
så analytics engine kan genbruge alle 103 tests.

Understøtter fleksibel kolonne-mapping: brugeren behøver ikke have præcise
kolonnenavne — parseren forsøger at auto-detektere baseret på almindelige navne.

Understøtter store filer (op til 2 GB):
- CSV: chunked parsing via pandas read_csv(chunksize=...)
- Excel: openpyxl read_only=True for streaming af store filer
- Automatisk valg: < 50 MB bruger standard metode, >= 50 MB bruger chunked/streaming
"""

import os
import pandas as pd
import re
from datetime import datetime
from typing import Optional, Callable
from openpyxl import load_workbook


# Størrelses-grænse for chunked parsing (50 MB)
LARGE_FILE_THRESHOLD = 50 * 1024 * 1024
CSV_CHUNK_SIZE = 10000


# Kolonnenavn-aliaser for auto-detektion
COLUMN_ALIASES = {
    "transaction_id": [
        "transaction_id", "transaktions_id", "trans_id", "voucher", "bilag",
        "bilagsnr", "bilagsnummer", "voucher_no", "doc_no", "document_number",
        "journal_entry", "posteringsnr", "entry_no", "id",
    ],
    "date": [
        "date", "dato", "transaction_date", "transaktionsdato", "bogføringsdato",
        "posting_date", "bogforingsdato", "document_date", "fakturadato",
        "invoice_date", "valuedato", "value_date",
    ],
    "account_id": [
        "account_id", "konto", "kontonr", "kontonummer", "account", "account_no",
        "account_number", "gl_account", "finans_konto", "hovedkonto",
    ],
    "account_description": [
        "account_description", "kontobeskrivelse", "kontonavn", "account_name",
        "account_desc", "beskrivelse_konto",
    ],
    "description": [
        "description", "beskrivelse", "text", "tekst", "narrative", "posting_text",
        "posteringstekst", "bilagstekst", "memo", "comment",
    ],
    "debit": [
        "debit", "debet", "debit_amount", "debitbeløb", "debitbelob",
    ],
    "credit": [
        "credit", "kredit", "credit_amount", "kreditbeløb", "kreditbelob",
    ],
    "amount": [
        "amount", "beløb", "belob", "total", "net_amount", "nettobeløb",
    ],
    "vat_amount": [
        "vat_amount", "moms", "momsbeløb", "momsbelob", "vat", "tax_amount",
        "momsbeloeb", "skat", "afgift",
    ],
    "vat_code": [
        "vat_code", "momskode", "tax_code", "moms_kode", "afgiftskode",
        "vat_type", "momstype", "tax_type",
    ],
    "vat_rate": [
        "vat_rate", "momssats", "momsprocent", "tax_rate", "moms_pct",
        "vat_pct", "vat_percentage",
    ],
    "supplier_id": [
        "supplier_id", "leverandør_id", "leverandornr", "leverandørnr",
        "vendor_id", "vendor_no", "supplier_no", "creditor_id", "kreditornr",
    ],
    "supplier_name": [
        "supplier_name", "leverandør", "leverandornavn", "leverandørnavn",
        "vendor_name", "vendor", "creditor_name",
    ],
    "customer_id": [
        "customer_id", "kunde_id", "kundenr", "customer_no", "debitor_id",
        "debitornr",
    ],
    "customer_name": [
        "customer_name", "kunde", "kundenavn", "customer", "debitor_name",
        "debitornavn",
    ],
    "invoice_number": [
        "invoice_number", "fakturanr", "fakturanummer", "invoice_no", "invoice",
        "ekstern_bilag", "external_doc",
    ],
    "currency": [
        "currency", "valuta", "valutakode", "currency_code",
    ],
    "journal_id": [
        "journal_id", "journal", "journalnr", "journal_no", "journal_type",
        "journaltype", "kladde",
    ],
    "period": [
        "period", "periode", "month", "måned", "maaned",
    ],
    "year": [
        "year", "år", "aar", "fiscal_year", "regnskabsår",
    ],
    "country": [
        "country", "land", "landekode", "country_code",
    ],
    "vat_number": [
        "vat_number", "momsnr", "momsnummer", "cvr", "cvr_nr", "cvrnummer",
        "tax_id", "vat_id", "vat_registration",
    ],
}


def _normalize_column_name(name):
    """Normalisér kolonnenavn til lowercase, strip whitespace og specialtegn."""
    return re.sub(r"[^a-z0-9_æøåü]", "", str(name).lower().strip().replace(" ", "_"))


def _detect_columns(df):
    """
    Auto-detektér hvilke kolonner der mapper til hvilke felter.
    Returnerer dict: standard_field_name -> actual_column_name
    """
    mapping = {}
    normalized = {_normalize_column_name(col): col for col in df.columns}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = _normalize_column_name(alias)
            if norm_alias in normalized:
                mapping[field] = normalized[norm_alias]
                break

    return mapping


def _detect_columns_from_names(column_names):
    """
    Auto-detektér kolonner fra en liste af kolonnenavne (bruges til streaming).
    Returnerer dict: standard_field_name -> actual_column_name
    """
    mapping = {}
    normalized = {_normalize_column_name(col): col for col in column_names}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = _normalize_column_name(alias)
            if norm_alias in normalized:
                mapping[field] = normalized[norm_alias]
                break

    return mapping


def _safe_float(value, default=0.0):
    """Konvertér en værdi til float sikkert."""
    if pd.isna(value) or value is None:
        return default
    try:
        # Håndtér danske talformater (1.234,56 -> 1234.56)
        s = str(value).strip()
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return default


def _safe_str(value, default=""):
    """Konvertér en værdi til string sikkert."""
    if pd.isna(value) or value is None:
        return default
    return str(value).strip()


def _safe_date(value):
    """Konvertér en værdi til ISO dato-streng."""
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    # Prøv at parse streng
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d"]:
        try:
            return datetime.strptime(str(value).strip()[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return str(value).strip()[:10]


def _get_row_value(row, col_map, field):
    """Hent værdi fra en row (dict eller tuple) via col_map."""
    col_name = col_map.get(field)
    if col_name is None:
        return None
    if isinstance(row, dict):
        return row.get(col_name)
    # For namedtuple from itertuples
    safe_col = col_name.replace(" ", "_")
    return getattr(row, safe_col, None) if hasattr(row, safe_col) else None


def _process_row(row, idx, col_map, is_tuple=False):
    """
    Processér en enkelt række til en transaktion + samle-data.
    Returnerer (txn, account_info, supplier_info, customer_info, vat_info)
    """
    if is_tuple:
        # For itertuples: row is a namedtuple, access via attribute
        def get_val(field):
            col_name = col_map.get(field)
            if col_name is None:
                return None
            # itertuples replaces spaces with underscores and prepends _ to numeric names
            try:
                return getattr(row, col_name, None)
            except AttributeError:
                return None
    else:
        # For dict rows (from chunked CSV or openpyxl)
        def get_val(field):
            col_name = col_map.get(field)
            if col_name is None:
                return None
            if isinstance(row, dict):
                return row.get(col_name)
            return row.get(col_name) if hasattr(row, 'get') else None

    # Bestem beløb (debit/credit eller samlet amount)
    debit = _safe_float(get_val("debit")) if "debit" in col_map else 0.0
    credit = _safe_float(get_val("credit")) if "credit" in col_map else 0.0

    if "amount" in col_map and debit == 0 and credit == 0:
        amount = _safe_float(get_val("amount"))
        if amount >= 0:
            debit = amount
        else:
            credit = abs(amount)

    # Transaktions-ID
    txn_id = _safe_str(get_val("transaction_id")) or f"ROW-{idx + 2}"

    # Dato
    date_val = get_val("date") if "date" in col_map else None
    date_str = _safe_date(date_val)

    # Konto
    account_id = _safe_str(get_val("account_id"))
    account_desc = _safe_str(get_val("account_description"))

    # Moms
    vat_amount = _safe_float(get_val("vat_amount")) if "vat_amount" in col_map else None
    vat_code = _safe_str(get_val("vat_code")) if "vat_code" in col_map else ""
    vat_rate = _safe_float(get_val("vat_rate")) if "vat_rate" in col_map else None

    # Leverandør / Kunde
    supplier_id = _safe_str(get_val("supplier_id")) if "supplier_id" in col_map else ""
    supplier_name = _safe_str(get_val("supplier_name")) if "supplier_name" in col_map else ""
    customer_id = _safe_str(get_val("customer_id")) if "customer_id" in col_map else ""
    customer_name = _safe_str(get_val("customer_name")) if "customer_name" in col_map else ""

    txn = {
        "transaction_id": txn_id,
        "date": date_str,
        "account_id": account_id,
        "account_description": account_desc,
        "description": _safe_str(get_val("description")),
        "debit_amount": debit,
        "credit_amount": credit,
        "vat_amount": vat_amount,
        "vat_code": vat_code,
        "vat_rate": vat_rate,
        "journal_id": _safe_str(get_val("journal_id")) or "IMPORT",
        "invoice_number": _safe_str(get_val("invoice_number")),
        "supplier_id": supplier_id,
        "supplier_name": supplier_name,
        "customer_id": customer_id,
        "customer_name": customer_name,
        "currency": _safe_str(get_val("currency")) or "DKK",
        "country": _safe_str(get_val("country")),
        "vat_number": _safe_str(get_val("vat_number")),
        "period": _safe_str(get_val("period")),
        "year": _safe_str(get_val("year")),
    }

    account_info = None
    if account_id:
        account_info = {
            "account_id": account_id,
            "description": account_desc,
            "account_type": "",
            "opening_balance": 0.0,
            "closing_balance": 0.0,
        }

    supplier_info = None
    if supplier_id:
        supplier_info = {
            "supplier_id": supplier_id,
            "name": supplier_name,
            "vat_number": _safe_str(get_val("vat_number")) if supplier_id else "",
            "country": _safe_str(get_val("country")),
        }

    customer_info = None
    if customer_id:
        customer_info = {
            "customer_id": customer_id,
            "name": customer_name,
            "vat_number": "",
            "country": "",
        }

    vat_info = None
    if vat_code:
        vat_info = {
            "tax_code": vat_code,
            "description": f"Momskode {vat_code}",
            "rate": vat_rate if vat_rate is not None else 0.0,
        }

    return txn, account_info, supplier_info, customer_info, vat_info


def _read_csv_detect_params(file_path):
    """Detektér CSV separator og encoding ved at læse en lille prøve."""
    for sep in [";", ",", "\t"]:
        for encoding in ["utf-8", "latin-1", "cp1252"]:
            try:
                df = pd.read_csv(file_path, sep=sep, encoding=encoding, nrows=5)
                if len(df.columns) > 1:
                    return sep, encoding
            except Exception:
                continue
    # Fallback
    return ",", "utf-8"


def _parse_csv_chunked(file_path, progress_callback=None):
    """
    Parsér en stor CSV fil i chunks af CSV_CHUNK_SIZE rækker.
    Bruger pandas read_csv med chunksize parameter.
    """
    sep, encoding = _read_csv_detect_params(file_path)

    # Først: tæl totale rækker for progress (hurtig scan)
    total_rows = 0
    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        for _ in f:
            total_rows += 1
    total_rows = max(total_rows - 1, 0)  # Minus header

    if total_rows == 0:
        return {
            "header": {},
            "accounts": [],
            "tax_table": [],
            "transactions": [],
            "suppliers": [],
            "customers": [],
            "parse_info": {"error": "Filen er tom", "rows": 0, "columns": 0},
        }

    transactions = []
    accounts_seen = {}
    suppliers_seen = {}
    customers_seen = {}
    vat_codes_seen = {}
    col_map = None
    columns_list = None
    rows_processed = 0

    reader = pd.read_csv(
        file_path, sep=sep, encoding=encoding, chunksize=CSV_CHUNK_SIZE
    )

    for chunk in reader:
        if col_map is None:
            col_map = _detect_columns(chunk)
            columns_list = list(chunk.columns)

        for row_tuple in chunk.itertuples(index=False):
            row_dict = {col: getattr(row_tuple, col, None)
                        for col in chunk.columns}
            txn, acct, supp, cust, vat = _process_row(
                row_dict, rows_processed, col_map, is_tuple=False
            )
            transactions.append(txn)

            if acct and acct["account_id"] not in accounts_seen:
                accounts_seen[acct["account_id"]] = acct
            if supp and supp["supplier_id"] not in suppliers_seen:
                suppliers_seen[supp["supplier_id"]] = supp
            if cust and cust["customer_id"] not in customers_seen:
                customers_seen[cust["customer_id"]] = cust
            if vat and vat["tax_code"] not in vat_codes_seen:
                vat_codes_seen[vat["tax_code"]] = vat

            rows_processed += 1

        if progress_callback and total_rows > 0:
            pct = min(int((rows_processed / total_rows) * 100), 100)
            progress_callback(pct, rows_processed, total_rows)

    parse_info = {
        "rows": rows_processed,
        "columns": len(columns_list) if columns_list else 0,
        "detected_columns": col_map or {},
        "unmapped_columns": [
            col for col in (columns_list or [])
            if col not in (col_map or {}).values()
        ],
        "source_type": "csv",
        "parsing_mode": "chunked",
    }

    return _build_result(transactions, accounts_seen, suppliers_seen,
                         customers_seen, vat_codes_seen, parse_info)


def _parse_excel_streaming(file_path, sheet_name=None, progress_callback=None):
    """
    Parsér en stor Excel fil med openpyxl read_only=True mode.
    Streamer rækker uden at loade hele filen i hukommelsen.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Læs header (første række)
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return {
            "header": {},
            "accounts": [],
            "tax_table": [],
            "transactions": [],
            "suppliers": [],
            "customers": [],
            "parse_info": {"error": "Filen er tom", "rows": 0, "columns": 0},
        }

    column_names = [str(h) if h is not None else f"col_{i}"
                    for i, h in enumerate(header_row)]
    col_map = _detect_columns_from_names(column_names)

    # Estimér total rækker (openpyxl max_row kan være upræcis i read_only)
    total_rows = ws.max_row - 1 if ws.max_row else 0

    transactions = []
    accounts_seen = {}
    suppliers_seen = {}
    customers_seen = {}
    vat_codes_seen = {}
    rows_processed = 0

    for row_values in rows_iter:
        row_dict = {column_names[i]: v for i, v in enumerate(row_values)
                    if i < len(column_names)}

        txn, acct, supp, cust, vat = _process_row(
            row_dict, rows_processed, col_map, is_tuple=False
        )
        transactions.append(txn)

        if acct and acct["account_id"] not in accounts_seen:
            accounts_seen[acct["account_id"]] = acct
        if supp and supp["supplier_id"] not in suppliers_seen:
            suppliers_seen[supp["supplier_id"]] = supp
        if cust and cust["customer_id"] not in customers_seen:
            customers_seen[cust["customer_id"]] = cust
        if vat and vat["tax_code"] not in vat_codes_seen:
            vat_codes_seen[vat["tax_code"]] = vat

        rows_processed += 1

        if progress_callback and rows_processed % 10000 == 0 and total_rows > 0:
            pct = min(int((rows_processed / total_rows) * 100), 100)
            progress_callback(pct, rows_processed, total_rows)

    wb.close()

    if progress_callback and total_rows > 0:
        progress_callback(100, rows_processed, total_rows)

    parse_info = {
        "rows": rows_processed,
        "columns": len(column_names),
        "detected_columns": col_map,
        "unmapped_columns": [
            col for col in column_names if col not in col_map.values()
        ],
        "source_type": "excel",
        "parsing_mode": "streaming",
    }

    return _build_result(transactions, accounts_seen, suppliers_seen,
                         customers_seen, vat_codes_seen, parse_info)


def _build_result(transactions, accounts_seen, suppliers_seen,
                  customers_seen, vat_codes_seen, parse_info):
    """Byg det endelige resultat-dict med header udledt fra data."""
    header = {
        "company_name": "",
        "registration_number": "",
        "currency": "DKK",
        "period_start": "",
        "period_end": "",
        "source": "Excel/CSV import",
    }

    dates = [t["date"] for t in transactions if t["date"]]
    if dates:
        sorted_dates = sorted(dates)
        header["period_start"] = sorted_dates[0]
        header["period_end"] = sorted_dates[-1]

    return {
        "header": header,
        "accounts": list(accounts_seen.values()),
        "tax_table": list(vat_codes_seen.values()),
        "transactions": transactions,
        "suppliers": list(suppliers_seen.values()),
        "customers": list(customers_seen.values()),
        "parse_info": parse_info,
    }


def parse_excel(file_path: str, sheet_name: Optional[str] = None,
                progress_callback: Optional[Callable] = None):
    """
    Parser en Excel/CSV fil og returnerer data i standardformatet.

    For store filer (>= 50 MB) bruges chunked/streaming parsing automatisk.
    progress_callback(percent, rows_done, total_rows) kaldes løbende for store filer.

    Returnerer samme struktur som SAF-T parseren:
    {
        "header": {...},
        "accounts": [...],
        "tax_table": [...],
        "transactions": [...],
        "suppliers": [...],
        "customers": [...],
        "parse_info": {...}  # Info om parsing-processen
    }
    """
    file_size = os.path.getsize(file_path)
    is_large = file_size >= LARGE_FILE_THRESHOLD
    is_csv = file_path.lower().endswith((".csv", ".tsv"))

    # Store filer: brug chunked/streaming parsing
    if is_large:
        if is_csv:
            return _parse_csv_chunked(file_path, progress_callback=progress_callback)
        else:
            return _parse_excel_streaming(
                file_path, sheet_name=sheet_name,
                progress_callback=progress_callback
            )

    # Små filer: brug standard metode (hurtigere for små filer)
    if is_csv:
        # Prøv forskellige separatorer og encodings
        df = None
        for sep in [";", ",", "\t"]:
            for encoding in ["utf-8", "latin-1", "cp1252"]:
                try:
                    df = pd.read_csv(file_path, sep=sep, encoding=encoding)
                    if len(df.columns) > 1:
                        break
                except Exception:
                    continue
            else:
                continue
            break
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)

    if df is None or df.empty:
        return {
            "header": {},
            "accounts": [],
            "tax_table": [],
            "transactions": [],
            "suppliers": [],
            "customers": [],
            "parse_info": {"error": "Filen er tom", "rows": 0, "columns": 0},
        }

    # Auto-detektér kolonner
    col_map = _detect_columns(df)

    parse_info = {
        "rows": len(df),
        "columns": len(df.columns),
        "detected_columns": col_map,
        "unmapped_columns": [
            col for col in df.columns
            if col not in col_map.values()
        ],
        "source_type": "csv" if is_csv else "excel",
        "parsing_mode": "standard",
    }

    # Byg transaktioner med itertuples() for bedre performance
    transactions = []
    accounts_seen = {}
    suppliers_seen = {}
    customers_seen = {}
    vat_codes_seen = {}

    total_rows = len(df)
    for idx, row_tuple in enumerate(df.itertuples(index=False)):
        # Konvertér til dict for ensartet processering
        row_dict = {col: getattr(row_tuple, col, None) if hasattr(row_tuple, col)
                    else row_tuple[i]
                    for i, col in enumerate(df.columns)}

        txn, acct, supp, cust, vat = _process_row(
            row_dict, idx, col_map, is_tuple=False
        )
        transactions.append(txn)

        if acct and acct["account_id"] not in accounts_seen:
            accounts_seen[acct["account_id"]] = acct
        if supp and supp["supplier_id"] not in suppliers_seen:
            suppliers_seen[supp["supplier_id"]] = supp
        if cust and cust["customer_id"] not in customers_seen:
            customers_seen[cust["customer_id"]] = cust
        if vat and vat["tax_code"] not in vat_codes_seen:
            vat_codes_seen[vat["tax_code"]] = vat

        if progress_callback and idx % 10000 == 0 and total_rows > 0:
            pct = min(int((idx / total_rows) * 100), 100)
            progress_callback(pct, idx, total_rows)

    if progress_callback:
        progress_callback(100, total_rows, total_rows)

    return _build_result(transactions, accounts_seen, suppliers_seen,
                         customers_seen, vat_codes_seen, parse_info)


def get_column_mapping_preview(file_path: str, sheet_name: Optional[str] = None):
    """
    Returnerer en preview af de første 5 rækker + auto-detekteret kolonne-mapping.
    Bruges til at lade brugeren bekræfte/rette mappingen før analyse.
    """
    if file_path.endswith(".csv"):
        for sep in [";", ",", "\t"]:
            for encoding in ["utf-8", "latin-1", "cp1252"]:
                try:
                    df = pd.read_csv(file_path, sep=sep, encoding=encoding, nrows=5)
                    if len(df.columns) > 1:
                        break
                except Exception:
                    continue
            else:
                continue
            break
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0, nrows=5)

    col_map = _detect_columns(df)

    return {
        "columns": list(df.columns),
        "preview_rows": df.head(5).fillna("").to_dict(orient="records"),
        "auto_mapping": col_map,
        "unmapped_columns": [col for col in df.columns if col not in col_map.values()],
        "required_fields": ["transaction_id", "date", "account_id", "amount/debit/credit"],
        "optional_fields": list(COLUMN_ALIASES.keys()),
    }
