#!/usr/bin/env python3
"""
build_traceability.py — auto-genererer sporbarhedsartefakterne fra regelkataloget:

  docs/VAT-Analytics_Regel-sporbarhedsmatrix.xlsx   (kontrol -> kilde -> modul -> test)
  docs/VAT-Analytics_Sporbarhedsrapport.md          (dækning + status)

Begge udledes fra catalog/rules.json, så de aldrig kommer ud af sync med koden.
Kør efter enhver katalogændring (efter tools/build_rules_catalog.py):

    python tools/build_traceability.py

`kilde` pr. kontrol tages fra kataloget (sat via catalog/rule_notes.json). Hvor
den er tom, vises kategoriens RETSOMRAADE som beskrivende område — de præcise
paragraffer pinnes af den fagansvarlige i rule_notes.json (felt "kilde").
"""

import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
_BACKEND = os.path.dirname(_HERE)
_REPO = os.path.dirname(_BACKEND)
_CATALOG = os.path.join(_BACKEND, "catalog", "rules.json")
_DOCS = os.path.join(_REPO, "docs")

# Retsområde pr. kategori (beskrivende — ikke opfundne paragraf-numre). Bruges som
# kilde hvor kataloget ikke har en præcis kilde. Den fagansvarlige pinner præcise
# momslov-paragraffer pr. kontrol i catalog/rule_notes.json.
RETSOMRAADE = {
    1: "Momsgrundlag og fradragsret (momsloven) — datakvalitet jf. Skattestyrelsens kontrolmetoder",
    2: "Dobbelt fradrag / dublethåndtering (momsloven, fradragsret)",
    3: "Momssatser: standardsats 25% og 0-sats (momsloven)",
    4: "Leveringssted, EU-handel og eksport: 0-sats og omvendt betalingspligt (momsloven)",
    5: "Leveringstidspunkt og angivelsesperiode — periodisering (momsloven)",
    6: "Gyldigt momsnummer og parts-/fakturakrav (momsloven, VIES)",
    7: "Beløbs- og tærskelkontroller (Skattestyrelsens kontrolmetoder)",
    8: "Statistisk anomalidetektion, Benford m.v. (Skattestyrelsens kontrolmetoder)",
    9: "Omvendt betalingspligt / reverse charge (momsloven)",
    10: "Opgørelse og afstemning af ind-/udgående moms; (delvis) fradragsret (momsloven)",
    11: "Svig/MTIC: solidarisk hæftelse og karruselindikatorer (momsloven, EU)",
    12: "E-handel, digitale ydelser og særordninger: OSS, brugtmoms, rejsebureau (momsloven)",
}

FONT = "Arial"
NAVY = "1B365D"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT = Font(name=FONT, size=10)
WRAP = Alignment(vertical="top", wrap_text=True)


def _load():
    with open(_CATALOG, encoding="utf-8") as f:
        return json.load(f)


def _kilde(rule):
    return rule.get("kilde") or RETSOMRAADE.get(rule["kategori_id"], "—")


def _join(v):
    return ", ".join(v) if isinstance(v, list) else (v or "")


def _style_sheet(ws, headers, widths):
    for col, (head, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=head)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _append(ws, values):
    ws.append(values)
    for c in ws[ws.max_row]:
        c.font = BODY_FONT
        c.alignment = WRAP


def _stats(cat):
    rules = cat["regler"]
    aktive = [r for r in rules if r["status"] == "aktiv"]
    inaktive = [r for r in rules if r["status"] != "aktiv"]
    med_kilde = [r for r in rules if r.get("kilde")]
    med_test = [r for r in rules if r.get("test")]
    return aktive, inaktive, med_kilde, med_test


def build_xlsx(cat):
    aktive, inaktive, med_kilde, med_test = _stats(cat)
    wb = Workbook()

    # --- Oversigt ---
    ws = wb.active
    ws.title = "Oversigt"
    _style_sheet(ws, ["Nøgletal", "Værdi"], [42, 64])
    facts = [
        ("Katalogversion", cat["catalog_version"]),
        ("Genereret", date.today().isoformat()),
        ("Kontroller i alt", cat["antal_kontroller"]),
        ("— aktive", len(aktive)),
        ("— inaktive (kræver kildedata / uden for scope)", len(inaktive)),
        ("Kategorier", len(cat["kategorier"])),
        ("Kontroller med præcis kilde udfyldt", f"{len(med_kilde)} / {len(cat['regler'])}"),
        ("Kontroller med dækkende test udfyldt", f"{len(med_test)} / {len(cat['regler'])}"),
        ("Inaktive test_id", ", ".join(str(r["test_id"]) for r in inaktive)),
    ]
    for k, v in facts:
        _append(ws, [k, v])

    # --- Sporbarhedsmatrix (kontrol -> kilde -> modul -> test) ---
    ws = wb.create_sheet("Sporbarhedsmatrix")
    _style_sheet(ws, ["ID", "Test", "Kontrol", "Kategori", "Severity", "Impact",
                      "Status", "Analyse-modul", "Default", "Retsområde / kilde",
                      "Kilde-modul", "Funktion", "Dækkende test", "Afhænger af"],
                 [10, 6, 40, 30, 14, 22, 22, 24, 10, 46, 40, 30, 40, 38])
    for r in cat["regler"]:
        _append(ws, [
            r["id"], r["test_id"], r["navn"], r["kategori"],
            _join(r["severity"]), _join(r["impact_type"]), r["status"],
            r.get("analyse_modul_navn", ""),
            "TIL" if r.get("default_aktiv") else "FRA",
            _kilde(r), r["modul"], r["funktion"],
            r.get("test") or "(udfyldes — Fase C)",
            r.get("afhaenger_af") or "",
        ])

    # --- Analyse-moduler ---
    if cat.get("analyse_moduler"):
        ws = wb.create_sheet("Analyse-moduler")
        _style_sheet(ws, ["Modul", "Nøgle", "Default", "Kontroller", "Beskrivelse"],
                     [24, 22, 10, 12, 70])
        for m in cat["analyse_moduler"]:
            _append(ws, [m["navn"], m["noegle"],
                         "TIL" if m["default_aktiv"] else "FRA",
                         m["antal_kontroller"], m["beskrivelse"]])

    # --- Kategorier ---
    ws = wb.create_sheet("Kategorier")
    _style_sheet(ws, ["ID", "Kategori", "Test-range", "Kontroller", "Aktive", "Inaktive"],
                 [6, 44, 14, 12, 10, 10])
    for c in cat["kategorier"]:
        lo, hi = c["test_range"]
        in_cat = [r for r in cat["regler"] if lo <= r["test_id"] <= hi]
        akt = sum(1 for r in in_cat if r["status"] == "aktiv")
        _append(ws, [c["id"], c["navn"], f"{lo}-{hi}", len(in_cat), akt, len(in_cat) - akt])

    os.makedirs(_DOCS, exist_ok=True)
    out = os.path.join(_DOCS, "VAT-Analytics_Regel-sporbarhedsmatrix.xlsx")
    wb.save(out)
    return out


def build_report(cat):
    aktive, inaktive, med_kilde, med_test = _stats(cat)
    L = []
    L.append("# VAT Analytics — Sporbarhedsrapport")
    L.append("")
    L.append(f"> Auto-genereret af `tools/build_traceability.py` · "
             f"katalogversion **{cat['catalog_version']}** · {date.today().isoformat()}")
    L.append("")
    L.append("## Dækning")
    L.append("")
    L.append(f"- Kontroller i alt: **{cat['antal_kontroller']}** "
             f"(aktive: **{len(aktive)}**, inaktive: **{len(inaktive)}**)")
    L.append(f"- Præcis kilde udfyldt: **{len(med_kilde)} / {len(cat['regler'])}** "
             f"(resten viser kategoriens retsområde indtil den fagansvarlige pinner paragraffen)")
    L.append(f"- Dækkende valideringstest udfyldt: **{len(med_test)} / {len(cat['regler'])}** "
             f"(valideringssuiten dækker alle {len(aktive)} aktive kontroller; de "
             f"{len(inaktive)} uden test er de inaktive)")
    L.append("")
    if cat.get("analyse_moduler"):
        L.append("## Analyse-moduler (momsrelevans-slankning)")
        L.append("")
        L.append("| Modul | Default | Kontroller | Beskrivelse |")
        L.append("|-------|---------|------------|-------------|")
        for m in cat["analyse_moduler"]:
            dflt = "TIL" if m["default_aktiv"] else "FRA"
            L.append(f"| {m['navn']} | {dflt} | {m['antal_kontroller']} | {m['beskrivelse']} |")
        L.append("")
    L.append("## Sporbarhedsmatrix (kontrol → analyse-modul → kilde → test)")
    L.append("")
    L.append("| ID | Kontrol | Status | Analyse-modul | Default | Retsområde / kilde | Test |")
    L.append("|----|---------|--------|---------------|---------|--------------------|------|")
    for r in cat["regler"]:
        test = r.get("test") or "(Fase C)"
        dflt = "TIL" if r.get("default_aktiv") else "FRA"
        L.append(f"| {r['id']} | {r['navn']} | {r['status']} | "
                 f"{r.get('analyse_modul_navn','')} | {dflt} | {_kilde(r)} | {test} |")
    L.append("")
    L.append("## Inaktive kontroller (beslutning og afhængighed)")
    L.append("")
    L.append("| ID | Kontrol | Afhænger af | Beslutning |")
    L.append("|----|---------|-------------|------------|")
    for r in inaktive:
        L.append(f"| {r['id']} | {r['navn']} | {r.get('afhaenger_af','')} | "
                 f"{r.get('scope_beslutning','')} |")
    L.append("")

    os.makedirs(_DOCS, exist_ok=True)
    out = os.path.join(_DOCS, "VAT-Analytics_Sporbarhedsrapport.md")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L) + "\n")
    return out


def main():
    cat = _load()
    xlsx = build_xlsx(cat)
    md = build_report(cat)
    print("Skrev", os.path.relpath(xlsx, _REPO))
    print("Skrev", os.path.relpath(md, _REPO))


if __name__ == "__main__":
    main()
